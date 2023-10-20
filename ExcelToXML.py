import os
from datetime import datetime
from xml.dom import minidom
import keyboard
from openpyxl import load_workbook
from colorama import Fore
import configparser
from xml.etree import ElementTree as ET

class PrintMsg:
    """
    Addition class - print message/Дополнительный класс красивой печати типовых сообщений
    """
    def __init__(self):
        self.IsPrintDebug: bool = False

    def print_service_message(self, value):
        print(f'{Fore.BLUE}{value}{Fore.WHITE}')

    def print_header(self, value):
        print(f'{Fore.YELLOW}{value}{Fore.WHITE}')

    def print_error(self, value):
        print(f'{Fore.RED}Error: {value}{Fore.WHITE}')

    def print_success(self, value):
        print(f'{Fore.GREEN}Success: {value}{Fore.WHITE}')

    def print_debug(self, value):
        if self.IsPrintDebug:
            print(f'{Fore.MAGENTA}{value}{Fore.WHITE}')


class AppSettings:
    """
    Addition class - settings/Дополнительный класс хранения настроек из CFG
    """
    def __init__(self):
        self.NumColumnHeadKey: int = 0
        self.IsPrintDebug: bool = False
        self.IsPrettyPrint: bool = False
        self.XML_ItemId: str = ''
        self.XML_ItemSpId: str = ''
        self.SkipEmptyAttr: str = ''

    def __str__(self):
        return f'AppSettings: {self.__dict__} '


def get_value_name_low(variable):
    """
    Get value/Получить значение настройки
    :param variable: Name like f"{appsettings.Server=}" => appsettings.Server
    :return: value
    """
    var_str = variable.split('=')[0].lower()
    return var_str


def get_class_value_name_low(variable):
    """
    Get name class.var to lower. Template: classname.valuename=/Получить имя настройки с нижнем регистре
    :param variable: Name like f"{appsettings.Server=}" => Server
    :return: lower case var name
    """
    var_str = get_value_name_low(variable)
    var_str = var_str.split('.')[1]
    return var_str


# Read config file/Чтение конфигурационного файла
def read_config(config_file_path):
    """
    Read exist file or create new whit default value/Прочитать или создать конфигурационный файл
    :param config_file_path: Config file full path
    """
    if os.path.exists(config_file_path):
        printmsg.print_header(f'Start ReadConfig')

        config = configparser.ConfigParser()
        config.read(config_file_path, "utf8")
        config.sections()

        var_settings_name = get_class_value_name_low(f"{appsettings.NumColumnHeadKey=}")
        appsettings.NumColumnHeadKey = config.has_option("Settings", var_settings_name) and config.getint("Settings",
                                                                                                          var_settings_name) or 0

        var_settings_name = get_class_value_name_low(f"{appsettings.XML_ItemId=}")
        appsettings.XML_ItemId = config.has_option("Settings", var_settings_name) and config.get("Settings",
                                                                                                 var_settings_name) or None
        # Note that the accepted values for the option are "1", "yes", "true", and "on"
        var_settings_name = get_class_value_name_low(f"{appsettings.IsPrintDebug=}")
        appsettings.IsPrintDebug = config.has_option("Settings", var_settings_name) and config.getboolean(
            "Settings",
            var_settings_name) or False

        # Note that the accepted values for the option are "1", "yes", "true", and "on"
        var_settings_name = get_class_value_name_low(f"{appsettings.IsPrettyPrint=}")
        appsettings.IsPrettyPrint = config.has_option("Settings", var_settings_name) and config.getboolean(
            "Settings",
            var_settings_name) or False

        var_settings_name = get_class_value_name_low(f"{appsettings.XML_ItemSpId=}")
        appsettings.XML_ItemSpId = config.has_option("Settings", var_settings_name) and config.get("Settings",
                                                                                                   var_settings_name) or None

        # Note that the accepted values for the option are "1", "yes", "true", and "on"
        var_settings_name = get_class_value_name_low(f"{appsettings.SkipEmptyAttr=}")
        appsettings.SkipEmptyAttr = config.has_option("Settings", var_settings_name) and config.getboolean(
            "Settings",
            var_settings_name) or False

        printmsg.print_success(f'Read config: {config_file_path}')
        return True
    else:
        printmsg.print_header(f'Start create config')
        config = configparser.ConfigParser()
        config.add_section("Settings")

        var_settings_name = get_class_value_name_low(f"{appsettings.NumColumnHeadKey=}")
        config.set("Settings", var_settings_name, '1')

        var_settings_name = get_class_value_name_low(f"{appsettings.IsPrintDebug=}")
        config.set("Settings", var_settings_name, 'false')

        var_settings_name = get_class_value_name_low(f"{appsettings.IsPrettyPrint=}")
        config.set("Settings", var_settings_name, 'false')

        var_settings_name = get_class_value_name_low(f"{appsettings.XML_ItemId=}")
        config.set("Settings", var_settings_name, 'order')

        var_settings_name = get_class_value_name_low(f"{appsettings.XML_ItemSpId=}")
        config.set("Settings", var_settings_name, 'ordersp01')

        var_settings_name = get_class_value_name_low(f"{appsettings.SkipEmptyAttr=}")
        config.set("Settings", var_settings_name, 'false')

        with open(config_file_path, "w") as config_file:
            config.write(config_file)

        printmsg.print_success(f'Create config: {config_file_path}')

        return False


def CreateXML(out_file_path):
    """
    Create result XML file/Создать результирующий файл XML
    :param out_file_path: Output file full path
    """
    printmsg.print_header(f"=======================================")
    printmsg.print_service_message(f"DEF WritingXML  - {out_file_path}")

    # Get key value
    key_column_name = list((i["name"] for i in list(DataHeader) if i["index"] == appsettings.NumColumnHeadKey))[0]
    # Get unique key
    head_list = list(set((i[key_column_name] for i in list(DataSetXLSX))))
    printmsg.print_debug(f'Key name: {key_column_name=}; Values = {list(head_list)}')

    # Only If exist Data
    if len(head_list) > 0:
        # XML: start write
        _xml_root = ET.Element("root")  # root

        _xml_Document = ET.SubElement(_xml_root, "fileinfo")  # info
        _xml_Document.set("version", "ExcelToXML")
        _xml_Document.set("datetime", f'{datetime.now()}')

        _xml_Document = ET.SubElement(_xml_root, "item")  # общее дерево
        _xml_Document.set("itemid", appsettings.XML_ItemId)

        # Header column name
        head_column_name_list = list((i for i in list(DataHeader) if i["is_sp"] == False))

        # SP column name
        sp_column_sp_list = list((i for i in list(DataHeader) if i["is_sp"] == True))

        # For unique head create SP
        for head_value in head_list:
            printmsg.print_debug(f"========{head_value}==========")
            sp_list = list((i for i in list(DataSetXLSX) if i[key_column_name] == head_value))
            if len(sp_list) > 0:
                head = sp_list[0]
                # Write head values
                _xml_Item = ET.SubElement(_xml_Document, "item")
                for head_name in head_column_name_list:
                    if (appsettings.SkipEmptyAttr and len(
                            head[head_name["name"]]) != 0) or not appsettings.SkipEmptyAttr:
                        _xml_Item.set(head_name["name"], head[head_name["name"]])
                        printmsg.print_debug(f'{head_name["name"]} = {head[head_name["name"]]}')

                # Write sp values
                _sp_exist = False
                for sp in sp_list:
                    for sp_name in sp_column_sp_list:
                        if len(sp[sp_name["name"]]) > 0:
                            _sp_exist = True
                            break
                # IF SP Exist
                if _sp_exist:
                    _xml_ItemSP = ET.SubElement(_xml_Item, "itemsp")
                    _xml_ItemSP.set("itemspid", appsettings.XML_ItemSpId)
                    for sp in sp_list:
                        printmsg.print_debug(f"========SP==========")
                        _xml_Item2 = ET.SubElement(_xml_ItemSP, "item")
                        for sp_name in sp_column_sp_list:
                            if (appsettings.SkipEmptyAttr and len(
                                    sp[sp_name["name"]]) != 0) or not appsettings.SkipEmptyAttr:
                                _xml_Item2.set(sp_name["name"], sp[sp_name["name"]])
                                printmsg.print_debug(f' {sp_name["name"]} = {sp[sp_name["name"]]}')

        # XML: end write
        _xml_tree = ET.ElementTree(_xml_root)  # записываем дерево в файл

        if appsettings.IsPrettyPrint:
            xmlstr = minidom.parseString(ET.tostring(_xml_root)).toprettyxml(indent="   ")

            with open(out_file_path, "w") as f:
                f.write(xmlstr)
        else:
            _xml_tree.write(out_file_path, encoding='utf-8', xml_declaration=True, method='xml')  # сохраняем файл
            _xml_tree.write(out_file_path, encoding='utf-8', xml_declaration=True, method='xml')  # сохраняем файл


def ParsingXLSX(in_file_path):
    """
    Read Excel file/Прочитать Excel файл
    :param in_file_path: In Excel file full path
    """
    printmsg.print_header(f"=======================================")
    printmsg.print_service_message(f"DEF ParsingXLSX - {in_file_path}")
    workbook = load_workbook(filename=in_file_path)
    sheet = workbook.active
    printmsg.print_service_message(f"File Count: row = {sheet.max_row}, column = {sheet.max_column}")

    printmsg.print_service_message(f"Key Column - {appsettings.NumColumnHeadKey}")

    # Read from space Row (empty appsettings.NumColumnHeadKey)
    first_in = True
    u_max_row = sheet.max_row
    for i in range(1, int(sheet.max_row)):
        if first_in and not sheet.cell(row=i, column=appsettings.NumColumnHeadKey).value:
            first_in = False
            u_max_row = i

    # Read from space Column (first Row, empty value)
    first_in = True
    u_max_col = sheet.max_column + 1
    for i in range(1, int(sheet.max_column)):
        if first_in and not sheet.cell(row=1, column=i).value:
            first_in = False
            u_max_col = i

    printmsg.print_service_message(f"User Count: row = {u_max_row}, column = {u_max_col}")

    # Read column name = attribute name
    for i in range(1, u_max_col):
        is_sp = str.lower(sheet.cell(row=1, column=i).value).strip().startswith("sp.")

        value = str.lower(sheet.cell(row=1, column=i).value).strip()

        if is_sp:
            value = value[3::]

        element = {"index": i, "is_sp": is_sp, "name": value}
        DataHeader.append(element)

    # Read dataset from file
    for i in range(2, u_max_row):
        row = {}
        for row_Head in DataHeader:
            row["Row"] = i
            value = sheet.cell(row=i, column=row_Head.get("index")).value
            if value is None:
                value = ""
            else:
                value = str(value).strip()
            row[row_Head.get("name")] = value
        DataSetXLSX.append(row)

    # DEBUG print Header
    printmsg.print_debug(f"\nDataHeader")
    for el in DataHeader:
        printmsg.print_debug(el)

    # DEBUG print DataSet
    printmsg.print_debug(f"\nDataSet")
    for el in DataSetXLSX:
        printmsg.print_debug(f"{el}")


def main():
    for in_file in os.listdir(os.getcwd()):
        if os.path.isfile(in_file) and in_file.endswith(".xlsx"):
            # Input file
            printmsg.print_header(f"Input: {in_file}")

            DataSetXLSX.clear()  # clear dataset
            DataHeader.clear()  # clear header

            # Parse file
            ParsingXLSX(os.path.join(os.getcwd(), in_file))

            # Create XML
            out_file = os.path.join(os.getcwd(), f"{os.path.splitext(in_file)[0]}.xml")
            CreateXML(out_file)


if __name__ == '__main__':
    printmsg = PrintMsg()
    appsettings = AppSettings()  # Настройки

    printmsg.print_header(f"Create: Cherepanov Maxim masygreen@gmail.com (c), 10.2023")

    currentDirectory = os.getcwd()
    configFilePath = os.path.join(currentDirectory, 'config.cfg')

    if read_config(configFilePath):
        printmsg.IsPrintDebug = appsettings.IsPrintDebug
        DataHeader = []
        DataSetXLSX = []
        DataSetProcess = []

        main()
    else:
        printmsg.print_error(f'Pleas edit default Config value: {configFilePath}')
        printmsg.print_service_message(f'Process skip...')

    printmsg.print_success(f"\n\n*All Process done.\n*Press Space to Exit ... It the longest shortcut \_(o0)_\...")
    keyboard.wait("space")